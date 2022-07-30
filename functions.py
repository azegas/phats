from openpyxl import Workbook, load_workbook
import pandas as pd
import os
from tkinter import *

# username = int(input("ivesk skaiciuka"))

def procedura():
    global username
    username = entry.get()  # gets entry text
    delete_cols_rows()
    no_sca()
    merge()
    convert_to_txt()
    istrink_bruksniukus()
    delete_trash()


def delete_cols_rows():
    wb = load_workbook('sourcefiles/' + str(username)+".xlsx")
    ws = wb['Sheet1']
    # print_rows()
    ws.delete_cols(1, 1)
    ws.delete_rows(1, 5)
    # print_rows()

    wb.save('output/' + str(username)+".xlsx")


def no_sca():
    df = pd.read_excel('output/' + str(username)+".xlsx", sheet_name='Sheet1')
    # print(df)
    # Replace sub string
    df = df.replace('SCA', '', regex=True)
    # print(df)
    datatoexcel = pd.ExcelWriter('output/' + str(username)+".xlsx")
    df.to_excel(datatoexcel)
    datatoexcel.save()


def merge():
    wb = load_workbook('output/' + str(username)+".xlsx")
    ws = wb['Sheet1']
    # print_rows()

    ws.delete_cols(1, 1)
    ws.delete_rows(1, 1)
    # print_rows()

    wb.save('output/' + str(username)+".xlsx")


def convert_to_txt():
    # exportas to .txt
    # https://www.youtube.com/watch?v=M5b669dAm2g
    xl = pd.ExcelFile('output/' + str(username)+".xlsx")
    xl.sheet_names

    for sheet in xl.sheet_names:
        file = pd.read_excel(xl, sheet_name=sheet)
        file.to_csv('output/' + str(username)+".txt", header=True, index=False)

    os.remove('output/' + str(username)+".xlsx")


def istrink_bruksniukus():
    file = open('output/' + str(username)+".txt")
    contents = file.read()
    # print(contents)
    replaced_contents = contents.replace(',', '_')
    # print(replaced_contents)

    with open('output/' + str(username)+".txt", "w") as f:
        for i in replaced_contents:
            f.write(i)


def delete_trash():
    # Karina
    # https://www.csnewbs.com/python-10c-remove-edit-lines

    with open('output/' + str(username)+".txt", 'r') as file:
        lines = file.readlines()
    with open('sourcefiles/' + str(username)+"_siuksles.txt", 'r') as file:
        to_delete = file.read().splitlines()

    file = open('output/' + str(username)+".txt", 'w')
    for line in lines:
        if line.rstrip() not in to_delete:
            file.write(line)
            # print(line, "liko po pravalymo")
    file.close()


def submit():
    username = entry.get()  # gets entry text
    print("paspausta")
    return username
    delete_cols_rows()
    # print("Hello "+username)
    # with open('readme.txt', 'w') as f:
    #     f.write(username)


def delete():
    entry.delete(0, END)  # deletes the line of text


def backspace():
    entry.delete(len(entry.get())-1, END) # delete last character
    

# inspiration - https://www.youtube.com/watch?v=VkTrrqnWjsg
window = Tk()

# title
window.title("PHATS .xlsx to .txt")

# label
label = Label(window, text="Šasiukas: ")
label.config(font=("Consolas", 30))
label.pack(side = LEFT)

# buttons
submit = Button(window, text="KONVERTUOTI", command=procedura)
submit.pack(side = RIGHT)

delete = Button(window, text="ištrinti viską", command=delete)
delete.pack(side = RIGHT)

backspace = Button(window, text="trinti simbolį", command=backspace)
backspace.pack(side = RIGHT)

entry = Entry()
entry.focus()
entry.config(font=('Ink Free', 50)) #change font
entry.config(bg='#111111') #background
entry.config(fg='#00FF00') #foreground
entry.config(width=10) #width displayed in characters
# entry.insert(0,'Spongebob') #set default text
#entry.config(state=DISABLED) #ACTIVE/DISABLED
# entry.config(show='*') #replace characters shown with x character
entry.pack()


window.mainloop()
