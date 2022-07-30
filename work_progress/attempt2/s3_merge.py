from openpyxl import Workbook, load_workbook
import pandas as pd

wb = load_workbook('output/3043210_s2_noSCA.xlsx')
ws = wb['Sheet1']
# print_rows()

ws.delete_cols(1, 1)
ws.delete_rows(1, 1)
# print_rows()

wb.save('output/3043210_s3_merge.xlsx')

# # cells to merge  
# ws.merge_cells('A1:B15')
# print_rows()
