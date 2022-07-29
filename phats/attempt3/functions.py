from openpyxl import Workbook, load_workbook
import pandas as pd
import os

skaiciukas = int(input("ivesk skaiciuka"))

def delete_cols_rows():
    wb = load_workbook('sourcefiles/' + str(skaiciukas)+".xlsx")
    ws = wb['Sheet1']
    # print_rows()
    ws.delete_cols(1, 1)
    ws.delete_rows(1, 5)
    # print_rows()

    wb.save('output/' + str(skaiciukas)+".xlsx")

def no_sca():
    df = pd.read_excel('output/' + str(skaiciukas)+".xlsx", sheet_name='Sheet1')
    # print(df)
    # Replace sub string
    df = df.replace('SCA', '', regex=True)
    # print(df)
    datatoexcel = pd.ExcelWriter('output/' + str(skaiciukas)+".xlsx")
    df.to_excel(datatoexcel)
    datatoexcel.save()


def merge():
    wb = load_workbook('output/' + str(skaiciukas)+".xlsx")
    ws = wb['Sheet1']
    # print_rows()

    ws.delete_cols(1, 1)
    ws.delete_rows(1, 1)
    # print_rows()

    wb.save('output/' + str(skaiciukas)+".xlsx")


def convert_to_txt():
    # exportas to .txt
    # https://www.youtube.com/watch?v=M5b669dAm2g
    xl = pd.ExcelFile('output/' + str(skaiciukas)+".xlsx")
    xl.sheet_names

    for sheet in xl.sheet_names:
        file = pd.read_excel(xl, sheet_name=sheet)
        file.to_csv('output/' + str(skaiciukas)+".txt", header=True, index=False)

    os.remove('output/' + str(skaiciukas)+".xlsx")


def istrink_bruksniukus():
    file = open('output/' + str(skaiciukas)+".txt")
    contents = file.read()
    # print(contents)
    replaced_contents = contents.replace(',', '_')
    # print(replaced_contents)

    with open('output/' + str(skaiciukas)+".txt", "w") as f:
        for i in replaced_contents:
            f.write(i)


def delete_trash():
    # Karina
    # https://www.csnewbs.com/python-10c-remove-edit-lines

    with open('output/' + str(skaiciukas)+".txt", 'r') as file:
        lines = file.readlines()
    with open('sourcefiles/' + str(skaiciukas)+"_siuksles.txt", 'r') as file:
        to_delete = file.read().splitlines()

    file = open('output/' + str(skaiciukas)+".txt", 'w')
    for line in lines:
        if line.rstrip() not in to_delete:
            file.write(line)
            # print(line, "liko po pravalymo")
    file.close()

delete_cols_rows()
no_sca()
merge()
convert_to_txt()
istrink_bruksniukus()
delete_trash()
