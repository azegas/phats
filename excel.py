from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# wb = load_workbook('grades.xlsx')
# ws = wb.active
# ws = wb['Grades']

# ws['A2'] = 'Antanas'

# wb.create_sheet("Test")

# print(wb.sheetnames)

# wb.save('grades.xlsx')

#############################################

# wb = Workbook()
# ws = wb.active
# ws.title = 'Data'

# ws.append(['Tim', 'Is', 'Great', '!'])
# ws.append(['Tim', 'Is', 'Great', '!'])
# ws.append(['Tim', 'Is', 'Great', '!'])
# ws.append(['end'])

# wb.save('tim.xlsx')

wb = load_workbook('tim.xlsx')
ws = wb['Grades']

ws.insert_rows(4)
ws.insert_rows(4)
ws.delete_rows(4)

ws.insert_cols(2)

# wb.create_sheet("sheet5")
# print(wb.sheetnames)
# ws = wb.active

# # check values
# for row in range(1, 11):
#     for col in range(1, 5):
#         # char = chr(65 + col)
#         char = get_column_letter(col)
#         print(ws[char + str(row)].value)
        
# # change values
# for row in range(1, 11):
#     for col in range(1, 5):
#         # char = chr(65 + col)
#         char = get_column_letter(col)
#         ws[char + str(row)] = char + str(row)

wb.save('tim.xlsx')
